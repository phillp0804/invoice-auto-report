"""報表彙整與 Excel 匯出服務。"""

from collections import defaultdict
from datetime import date as date_cls
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from models.invoice import Invoice
from schemas.report_schema import (
    CategorySummary,
    DashboardResponse,
    DepartmentSummary,
    EmployeeSummary,
)

# 儀表板/報表只統計總務已確認的發票，待審核與已退回不計入財務彙整
_CONFIRMED_STATUS = "已確認"
_PENDING_STATUS = "待審核"

_GROUP_BY_OPTIONS = ("department", "employee", "category")

_AMOUNT_FORMAT = "#,##0"
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_TOTAL_FONT = Font(bold=True)
_TOTAL_BORDER = Border(top=Side(style="thin"))


class ReportService:
    """報表彙整與匯出服務。"""

    def get_dashboard_data(
        self,
        db: Session,
        year: int | None = None,
        month: int | None = None,
    ) -> DashboardResponse:
        """取得儀表板彙整資料。

        Args:
            db: 資料庫 session。
            year: 篩選年份（可選）。
            month: 篩選月份（可選，須搭配 year 使用）。

        Returns:
            依類別/員工/部門彙整的儀表板資料。
        """
        invoices = self._confirmed_invoices_query(db, year, month).all()
        pending_count = (
            db.query(Invoice).filter(Invoice.status == _PENDING_STATUS).count()
        )

        category_totals: dict[str, dict] = defaultdict(
            lambda: {"total_amount": Decimal("0"), "count": 0}
        )
        employee_totals: dict[int, dict] = defaultdict(
            lambda: {"total_amount": Decimal("0"), "count": 0, "user_name": ""}
        )
        department_totals: dict[int, dict] = defaultdict(
            lambda: {"total_amount": Decimal("0"), "count": 0, "department_name": ""}
        )

        total_amount = Decimal("0")
        for invoice in invoices:
            amount = invoice.amount or Decimal("0")
            total_amount += amount

            category_key = invoice.category or "未分類"
            category_totals[category_key]["total_amount"] += amount
            category_totals[category_key]["count"] += 1

            employee_totals[invoice.user_id]["total_amount"] += amount
            employee_totals[invoice.user_id]["count"] += 1
            employee_totals[invoice.user_id]["user_name"] = invoice.user.name

            department = invoice.user.department
            if department is not None:
                department_totals[department.id]["total_amount"] += amount
                department_totals[department.id]["count"] += 1
                department_totals[department.id]["department_name"] = department.name

        return DashboardResponse(
            total_amount=total_amount,
            total_count=len(invoices),
            pending_count=pending_count,
            by_category=[
                CategorySummary(category=key, **data)
                for key, data in category_totals.items()
            ],
            by_employee=[
                EmployeeSummary(user_id=key, **data)
                for key, data in employee_totals.items()
            ],
            by_department=[
                DepartmentSummary(department_id=key, **data)
                for key, data in department_totals.items()
            ],
        )

    def export_excel(
        self,
        db: Session,
        year: int,
        month: int,
        group_by: str = "department",
    ) -> BytesIO:
        """匯出 Excel 報表。

        Args:
            db: 資料庫 session。
            year: 報表年份。
            month: 報表月份。
            group_by: 分組方式（"department" / "employee" / "category"）。

        Returns:
            Excel 檔案的 BytesIO 物件（游標已重置至開頭，可直接讀取）。

        Raises:
            ValueError: group_by 不在允許的分組方式內。
        """
        if group_by not in _GROUP_BY_OPTIONS:
            raise ValueError(
                f"不支援的分組方式：{group_by}，須為 department / employee / category"
            )

        dashboard = self.get_dashboard_data(db, year=year, month=month)

        column_label, summaries, name_getter = {
            "department": ("部門", dashboard.by_department, lambda s: s.department_name),
            "employee": ("員工", dashboard.by_employee, lambda s: s.user_name),
            "category": ("類別", dashboard.by_category, lambda s: s.category),
        }[group_by]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{year}-{month:02d} 報表"

        sheet.append([column_label, "金額", "件數"])
        for cell in sheet[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
        sheet.freeze_panes = "A2"

        for summary in summaries:
            sheet.append([name_getter(summary), float(summary.total_amount), summary.count])

        total_row = sheet.max_row + 1
        sheet.append(["總計", float(dashboard.total_amount), dashboard.total_count])
        for cell in sheet[total_row]:
            cell.font = _TOTAL_FONT
            cell.border = _TOTAL_BORDER

        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].number_format = _AMOUNT_FORMAT

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 10

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _confirmed_invoices_query(
        self, db: Session, year: int | None, month: int | None
    ):
        """建立已確認發票的查詢，選擇性依年份/月份篩選。"""
        query = db.query(Invoice).filter(Invoice.status == _CONFIRMED_STATUS)
        if year is not None:
            start = date_cls(year, month or 1, 1)
            if month is not None:
                end = (
                    date_cls(year + 1, 1, 1)
                    if month == 12
                    else date_cls(year, month + 1, 1)
                )
            else:
                end = date_cls(year + 1, 1, 1)
            query = query.filter(Invoice.date >= start, Invoice.date < end)
        return query
